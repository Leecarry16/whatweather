## Written for classficate the image data and rewrite Excel data
## Contact by kesm03122@gmail.com

from openpyxl import Workbook, load_workbook
from importlib.resources import path

import pandas as pd
import openpyxl
import hashlib
import shutil
import sys
import re
import os

process = "C:/Users/user/Desktop/School/AI_Build_Project_test/Term Projecct/Program/이미지 분류 진행과정.xlsx"
im_path = "C:/Users/user/Desktop/School/AI_Build_Project_test/Term Projecct/Program/rename"
file_list = os.listdir(im_path)
rename_path = "C:/Users/user/Desktop\School/AI_Build_Project_test/Term Projecct/Program/rename/{}"

pm_dir = ["0105", "0610", "1115", "1620", "2125", "2630", "3135", "3640", "4145", "4650", "5155", "5660", "6165", "6670", "7175", "7680", "8185", "8690", "9195", "9600", "nan", "error"]

pm_path = "C:/Users/user/Desktop/School/AI_Build_Project_test/Term Projecct/Program/PM10 Class/{}"

#print(dataframeXlsx)

def get_files_count(folder_path):
	dirListing = os.listdir(folder_path)
	return len(dirListing)

sum_before = 0
for pm_dir_num in range(22):
	print(pm_dir[pm_dir_num], get_files_count(pm_path.format(pm_dir[pm_dir_num])))
	sum_before = sum_before + int(get_files_count(pm_path.format(pm_dir[pm_dir_num])))

print("분류 전 폴더 내 파일 개수: ", sum_before)

pm_exc = pd.read_excel("C:/Users/user/Desktop/School/AI_Build_Project_test/Term Projecct/Program/PM10_20220522.xlsx")

num = 0
test_list = []

for file in file_list:
	name = file.split(',')[0]
	name_y = name[0:4]
	name_m = name[4:6]
	name_d = name[6:8]
	name_h = name[8:10]
	name_ymd = name_y + "." + name_m + '.' + name_d + '.'
	for row_num in range(2, 33):
		if name_ymd in str(pm_exc.iat[row_num, 0]):
			pm_date = pm_exc.iat[row_num, 1]
			for col_num in range(1, 25):
				if name_h in re.sub(r'[^0-9]', '', str(pm_exc.iat[1, col_num])):
					pm_hour = pm_exc.iat[2, col_num]
					pm_cell = pm_exc.iat[row_num, col_num]
					test_list.append(pm_cell)
					try:
						for pm_cls in range(20):
							if (1 + (pm_cls*5)) <= pm_cell <= (5 + (pm_cls * 5)):
								shutil.move(rename_path.format(file_list[num]), pm_path.format(pm_dir[pm_cls]))
							elif pm_cell == float('nan'):
								shutil.move(rename_path.format(file_list[num]), pm_path.format("nan"))
					except shutil.Error:
						try:
							shutil.move(rename_path.format(file_list[num]), pm_path.format("error"))
						except shutil.Error:
							os.remove(rename_path.format(file_list[num]))

					num = num +1

print("분류한 파일 개수: ", num)

#print(pm_exc)

rb = load_workbook(process)
ws = rb['Sheet1']
wb = openpyxl.load_workbook(process)

Sheet1 = wb.active 

sum_after = 0

for pm_dir_num in range(22):
	Sheet1['B{}'.format(pm_dir_num + 2)] = int(get_files_count(pm_path.format(pm_dir[pm_dir_num])))
	sum_after = sum_after + int(get_files_count(pm_path.format(pm_dir[pm_dir_num])))

Sheet1['B24'] = sum_after
print("분류 후 폴더 내 파일 개수: ",sum_after)

sum_before = 0
for pm_dir_num in range(22):
	print(pm_dir[pm_dir_num], get_files_count(pm_path.format(pm_dir[pm_dir_num])))
	sum_before = sum_before + int(get_files_count(pm_path.format(pm_dir[pm_dir_num])))

wb.save(process)



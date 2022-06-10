import os
from sklearn.model_selection import train_test_split
import shutil
from openpyxl import Workbook, load_workbook
import openpyxl

test_path = "C:/Users/user/Desktop/School/AI_Build_Project_test/Term Projecct/Program/Code/dataset/test/{}"
train_path = "C:/Users/user/Desktop\School/AI_Build_Project_test/Term Projecct/Program/Code/dataset/train/{}"
process = "C:/Users/user/Desktop/School/AI_Build_Project_test/Term Projecct/Program/Code/dataset/data_set.xlsx"
pm_dir = ["0105", "0610", "1115", "1620", "2125", "2630", "3135", "3640", "4145", "4650", "5155", "5660", "6165", "6670", "7175", "7680", "8185", "8690", "9195", "9600", "error"]

def get_files_count(folder_path):
	dirListing = os.listdir(folder_path)
	return len(dirListing)

for pm_class in range(0,20):
    file_ls = os.listdir(f'C:/Users/user/Desktop/School/AI_Build_Project_test/Term Projecct/Program/PM10 Class/{pm_dir[pm_class]}')
    if len(file_ls)  < 3:
        print(pm_dir[pm_class], "--파일 부족--")
    
    else:
        train_ls, test_ls = train_test_split(file_ls, test_size=0.3, shuffle=True, random_state=1234)
        for train in range(len(train_ls)):
            shutil.copy(file_ls + '/' + train_ls[train], train_path.format(pm_dir[pm_class]))
        for test in range(len(test_ls)):
            shutil.copy(file_ls + '/' + test_ls[train], test_path.format(pm_dir[pm_class]))
        
        print()
        print(pm_dir[pm_class])
        print("Train: ",len(train_ls))
        print("Test: ", len(test_ls))

        '''print(train_ls[:10])
        print(test_ls[:10])'''

rb = load_workbook(process)
ws = rb['Sheet1']
wb = openpyxl.load_workbook(process)

Sheet1 = wb.active 

for pm_dir_num in range(20):
	Sheet1['E{}'.format(pm_dir_num + 2)] = int(get_files_count(train_path.format(pm_dir[pm_dir_num])))
	train_sum = train_sum + int(get_files_count(train_path.format(pm_dir[pm_dir_num])))
Sheet1['E22'] = train_sum

for pm_dir_num in range(20):
	Sheet1['B{}'.format(pm_dir_num + 2)] = int(get_files_count(test_path.format(pm_dir[pm_dir_num])))
	test_sum = test_sum + int(get_files_count(test_path.format(pm_dir[pm_dir_num])))
Sheet1['B22'] = test_sum

wb.save(process)

'''
print(len(file_ls))

train_ls, test_ls = train_test_split(file_ls, test_size=0.3, shuffle=True, random_state=1234)

print(len(train_ls))
print(len(test_ls)) 

print(train_ls[:10])
print(test_ls[:10])
'''
